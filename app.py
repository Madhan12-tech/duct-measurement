from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import sqlite3
from datetime import datetime
import pandas as pd
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

app = Flask(__name__)
app.secret_key = 'secretkey'

def init_db():
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT,
            enquiry_no TEXT,
            office_no TEXT,
            site_engineer TEXT,
            site_contact TEXT,
            location TEXT,
            timestamp DATETIME
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS duct_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            duct_no TEXT,
            duct_type TEXT,
            width1 REAL,
            height1 REAL,
            width2 REAL,
            height2 REAL,
            length_or_radius REAL,
            quantity INTEGER,
            degree_or_offset REAL,
            factor REAL DEFAULT 1.0,
            gauge TEXT,
            area REAL,
            nuts_bolts INTEGER,
            cleat INTEGER,
            gasket REAL,
            corner_pieces INTEGER,
            timestamp DATETIME
        )
    ''')
    conn.commit()
    conn.close()

init_db()

@app.route('/')
def index():
    conn = sqlite3.connect('data.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM projects ORDER BY id DESC")
    projects = cursor.fetchall()
    conn.close()
    return render_template('home.html', projects=projects)

@app.route('/save_project', methods=['POST'])
def save_project():
    form = request.form
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('''
        INSERT INTO projects (project_name, enquiry_no, office_no, site_engineer, site_contact, location, timestamp)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        form['project_name'], form['enquiry_no'], form['office_no'],
        form['site_engineer'], form['site_contact'], form['location'], datetime.now()
    ))
    conn.commit()
    project_id = c.lastrowid
    conn.close()
    return redirect(url_for('home', project_id=project_id))

@app.route('/home/<int:project_id>')
def home(project_id):
    conn = sqlite3.connect('data.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM projects WHERE id=?", (project_id,))
    project = cursor.fetchone()
    cursor.execute("SELECT * FROM duct_entries WHERE project_id=? ORDER BY id DESC", (project_id,))
    entries = cursor.fetchall()
    conn.close()

    def safe_total(index): return sum(float(e[index]) for e in entries if e[index] is not None)
    def area_by_gauge(g): return sum(e[13] for e in entries if e[12] == g)

    return render_template('duct_entry.html',
        project=project,
        entries=entries,
        project_id=project_id,
        total_qty=safe_total(9),
        total_area=safe_total(13),
        total_bolts=safe_total(14),
        total_cleat=safe_total(15),
        total_gasket=safe_total(16),
        total_corner=safe_total(17),
        area_24g=area_by_gauge('24g'),
        area_22g=area_by_gauge('22g'),
        area_20g=area_by_gauge('20g'),
        area_18g=area_by_gauge('18g')
    )

@app.route('/add_duct', methods=['POST'])
def add_duct():
    form = request.form
    id_ = form.get('id')

    try:
        project_id = int(form['project_id'])
        duct_no = form['duct_no']
        duct_type = form['duct_type']
        factor = float(form.get('factor') or 1.0) if duct_type in ['RED', 'OFFSET', 'SHOE', 'ELB'] else 1.0

        width1 = float(form.get('width1') or 0)
        height1 = float(form.get('height1') or 0)
        width2 = float(form.get('width2') or 0)
        height2 = float(form.get('height2') or 0)
        length = float(form.get('length_or_radius') or 0)
        quantity = int(form.get('quantity') or 0)
        degree = float(form.get('degree_or_offset') or 0)

        if not duct_no or width1 == 0 or height1 == 0 or length == 0 or quantity == 0:
            flash("Please fill in all required fields.")
            return redirect(url_for('home', project_id=project_id))

    except Exception as e:
        flash(f"Invalid input: {e}")
        return redirect(url_for('home', project_id=form.get('project_id')))

    # Gauge logic
    if width1 <= 375 and height1 <= 375:
        gauge = '24g'
    elif width1 <= 600 and height1 <= 600:
        gauge = '22g'
    elif width1 <= 900 and height1 <= 900:
        gauge = '20g'
    else:
        gauge = '18g'

    # Area calculation
    if duct_type == 'ST':
        area = 2 * (width1 + height1) / 1000 * (length / 1000) * quantity
    elif duct_type == 'RED':
        area = (width1 + height1 + width2 + height2) / 1000 * (length / 1000) * quantity * factor
    elif duct_type == 'DUM':
        area = (width1 * height1) / 1_000_000 * quantity
    elif duct_type == 'OFFSET':
        area = (width1 + height1 + width2 + height2) / 1000 * ((length + degree) / 1000) * quantity * factor
    elif duct_type == 'SHOE':
        area = (width1 + height1) * 2 / 1000 * (length / 1000) * quantity * factor
    elif duct_type == 'VANES':
        area = width1 / 1000 * (2 * 3.14 * (width1 / 1000) / 2) / 4 * quantity
    elif duct_type == 'ELB':
        area = 2 * (width1 + height1) / 1000 * ((height1 / 2 / 1000) + (length / 1000) * (3.14 * (degree / 180))) * quantity * factor
    else:
        area = 0

    nuts_bolts = quantity * 4
    cleat = quantity * (4 if gauge == '24g' else 8 if gauge == '22g' else 10 if gauge == '20g' else 12)
    gasket = (width1 + height1 + width2 + height2) / 1000 * quantity
    corner = 0 if duct_type == 'DUM' else quantity * 8

    conn = sqlite3.connect('data.db')
    c = conn.cursor()

    if id_:
        c.execute('''
            UPDATE duct_entries SET project_id=?, duct_no=?, duct_type=?, width1=?, height1=?, width2=?, height2=?,
            length_or_radius=?, quantity=?, degree_or_offset=?, factor=?, gauge=?, area=?, nuts_bolts=?, cleat=?,
            gasket=?, corner_pieces=?, timestamp=? WHERE id=?
        ''', (project_id, duct_no, duct_type, width1, height1, width2, height2,
              length, quantity, degree, factor, gauge, area, nuts_bolts, cleat, gasket, corner, datetime.now(), id_))
        flash("Duct updated")
    else:
        c.execute('''
            INSERT INTO duct_entries (
                project_id, duct_no, duct_type, width1, height1, width2, height2,
                length_or_radius, quantity, degree_or_offset, factor, gauge, area,
                nuts_bolts, cleat, gasket, corner_pieces, timestamp
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            project_id, duct_no, duct_type, width1, height1, width2, height2,
            length, quantity, degree, factor, gauge, area,
            nuts_bolts, cleat, gasket, corner, datetime.now()
        ))
        flash("Duct added")

    conn.commit()
    conn.close()
    return redirect(url_for('home', project_id=project_id))

@app.route('/edit/<int:id>')
def edit_duct(id):
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute("SELECT * FROM duct_entries WHERE id=?", (id,))
    entry = c.fetchone()
    if not entry:
        flash("Entry not found")
        return redirect(url_for('index'))
    project_id = entry[1]
    c.execute("SELECT * FROM projects WHERE id=?", (project_id,))
    project = c.fetchone()
    c.execute("SELECT * FROM duct_entries WHERE project_id=? ORDER BY id DESC", (project_id,))
    entries = c.fetchall()
    conn.close()

    def total(i): return sum(float(e[i]) for e in entries if e[i])
    def area_g(g): return sum(e[13] for e in entries if e[12] == g)

    return render_template('duct_entry.html', edit_entry=entry, project=project, entries=entries,
        project_id=project_id,
        total_qty=total(9),
        total_area=total(13),
        total_bolts=total(14),
        total_cleat=total(15),
        total_gasket=total(16),
        total_corner=total(17),
        area_24g=area_g('24g'),
        area_22g=area_g('22g'),
        area_20g=area_g('20g'),
        area_18g=area_g('18g')
    )

@app.route('/delete/<int:id>')
def delete_duct(id):
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute("SELECT project_id FROM duct_entries WHERE id=?", (id,))
    project_id = c.fetchone()[0]
    c.execute("DELETE FROM duct_entries WHERE id=?", (id,))
    conn.commit()
    conn.close()
    flash("Deleted successfully")
    return redirect(url_for('home', project_id=project_id))

@app.route('/export_excel/<int:project_id>')
def export_excel(project_id):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font

    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute("SELECT * FROM projects WHERE id=?", (project_id,))
    project = c.fetchone()
    c.execute("SELECT * FROM duct_entries WHERE project_id=?", (project_id,))
    entries = c.fetchall()
    conn.close()

    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duct Entries"

    # Project Info
    ws['A1'] = "Project:"
    ws['B1'] = project[1]
    ws['A2'] = "Enquiry No:"
    ws['B2'] = project[2]
    ws['C2'] = "Office No:"
    ws['D2'] = project[3]
    ws['A3'] = "Site Engineer:"
    ws['B3'] = project[4]
    ws['C3'] = "Site Contact:"
    ws['D3'] = project[5]
    ws['A4'] = "Location:"
    ws['B4'] = project[6]

    # Headers
    headers = ['#', 'Duct No', 'Type', 'W1', 'H1', 'W2', 'H2', 'Len/Rad', 'Qty',
               'Deg/Off', 'Factor', 'Gauge', 'Area', 'Nuts', 'Cleat', 'Gasket', 'Corner']
    ws.append([])  # Row 5: empty
    header_row = 6
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Data Rows
    start_row = header_row + 1
    for idx, entry in enumerate(entries, start=1):
        row = [idx, entry[2], entry[3], entry[4], entry[5], entry[6], entry[7],
               entry[8], entry[9], entry[10], entry[11], entry[12], entry[13],
               entry[14], entry[15], entry[16], entry[17]]
        for col_num, val in enumerate(row, 1):
            ws.cell(row=start_row + idx - 1, column=col_num, value=val)

    # Totals Row
    total_row = start_row + len(entries)
    ws.cell(row=total_row, column=1, value="Total")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
    ws.cell(row=total_row, column=9, value=sum(e[9] for e in entries))    # Qty
    ws.cell(row=total_row, column=13, value=sum(e[13] for e in entries))  # Area
    ws.cell(row=total_row, column=14, value=sum(e[14] for e in entries))  # Nuts
    ws.cell(row=total_row, column=15, value=sum(e[15] for e in entries))  # Cleat
    ws.cell(row=total_row, column=16, value=sum(e[16] for e in entries))  # Gasket
    ws.cell(row=total_row, column=17, value=sum(e[17] for e in entries))  # Corner

    # Auto-size
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="duct_entries_with_totals.xlsx")


@app.route('/export_pdf/<int:project_id>')
def export_pdf(project_id):
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet

    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute("SELECT * FROM projects WHERE id=?", (project_id,))
    project = c.fetchone()
    c.execute("SELECT * FROM duct_entries WHERE project_id=?", (project_id,))
    entries = c.fetchall()
    conn.close()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Project Info
    project_info = f"""
    <b>Project:</b> {project[1]}<br/>
    <b>Enquiry No:</b> {project[2]} &nbsp;&nbsp;&nbsp; <b>Office No:</b> {project[3]}<br/>
    <b>Site Engineer:</b> {project[4]} &nbsp;&nbsp;&nbsp; <b>Contact:</b> {project[5]}<br/>
    <b>Location:</b> {project[6]}
    """
    elements.append(Paragraph(project_info, styles['Normal']))
    elements.append(Spacer(1, 12))

    # Table Headers
    headers = ['#', 'Duct No', 'Type', 'W1', 'H1', 'W2', 'H2', 'Len/Rad', 'Qty',
               'Deg/Off', 'Factor', 'Gauge', 'Area', 'Nuts', 'Cleat', 'Gasket', 'Corner']
    data = [headers]

    # Data rows
    for idx, entry in enumerate(entries, start=1):
        row = [
            idx, entry[2], entry[3], entry[4], entry[5], entry[6], entry[7],
            entry[8], entry[9], entry[10], entry[11], entry[12], f"{entry[13]:.2f}",
            entry[14], entry[15], f"{entry[16]:.2f}", entry[17]
        ]
        data.append(row)

    # Totals Row
    total_row = [
        "Total", '', '', '', '', '', '', '', 
        sum(e[9] for e in entries), '', '', '',
        f"{sum(e[13] for e in entries):.2f}",
        sum(e[14] for e in entries),
        sum(e[15] for e in entries),
        f"{sum(e[16] for e in entries):.2f}",
        sum(e[17] for e in entries)
    ]
    data.append(total_row)

    # Table Style
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold')
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='duct_entries_with_totals.pdf')


if __name__ == '__main__':
    app.run(debug=True)
